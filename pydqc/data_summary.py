import pandas as pd 
import numpy as np 
import os
import shutil

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.formatting.rule import DataBar, FormatObject, Rule

import datetime
from sklearn.externals.joblib import Parallel, delayed

import matplotlib.pyplot as plt

import seaborn as sns 
sns.set_style('white')

from .dqc_utils import (
    _style_range, _get_scale_draw_values, _draw_texts,
    _adjust_ws, _insert_df, _insert_numeric_results
)

import warnings
warnings.filterwarnings('ignore')

# global color values
VER_LINE = "#4BACC6"
TEXT_LIGHT = "#DAEEF3"
DIS_LINE = "#F79646"


def distribution_summary_pretty(_value_df, col, figsize=None, date_flag=False):
    """
    Draw pretty distribution graph for a single column

    Parameters
    ----------
    _value_df: pandas dataframe
        the dataframe that contains the values of 'col'
    col: name of the column
    figsize: (width, height)
        Size of the figure
    date_flag: boolean
        Whether the column is date type
    """

    # colors for graph
    DIS_LINE = "#F79646"

    # copy the raw dataframe
    value_df = _value_df.copy()

    if date_flag:
        numeric_col = '%s_numeric' %(col)
        if numeric_col not in value_df.columns.values:
            snapshot_date_now = str(datetime.datetime.now().date())
            value_df[numeric_col] = (pd.to_datetime(snapshot_date_now) - pd.to_datetime(value_df[col],
                errors='coerce')).astype('timedelta64[M]', errors='ignore')
    else:
        numeric_col = col

    # get min, mean, median, max
    value_min = value_df[numeric_col].min()
    value_mean = value_df[numeric_col].mean()
    value_median = value_df[numeric_col].median()
    value_max = value_df[numeric_col].max()

    if date_flag:
        date_min = pd.to_datetime(value_df[col], errors='coerce').min()
        date_max = pd.to_datetime(value_df[col], errors='coerce').max()

    num_uni = value_df[col].dropna().nunique()
    value_dropna = value_df[numeric_col].dropna().values

    # get distribution
    scale_flg = 0
    draw_values = value_dropna
    draw_value_4 = [value_min, value_mean, value_median, value_max]
    if np.max([abs(value_min), abs(value_max)]) >= pow(10, 6):
        scale_flg = 1
        draw_values, draw_value_4 = _get_scale_draw_values(draw_values, draw_value_4)

    # draw and save distribution graph
    plt.clf()
    if figsize is not None:
        plt.figure(figsize)
    else:
        plt.figure(figsize=(10, 6))

    if scale_flg:
        plt.title('%s (log10 scale)' %(col))
    else:
        plt.title('%s' %(col))

    # if unique level is less than 10, draw countplot instead
    if num_uni <= 10:
        temp_df = pd.DataFrame(draw_values, columns=['value'])
        sns.countplot(temp_df['value'], color=DIS_LINE)
        if num_uni > 5:
            plt.xticks(rotation=90)
    else:
        ax = sns.distplot(draw_values, color=DIS_LINE, norm_hist=True, hist=False)
        y_low, y_up = ax.get_ylim()

        if date_flag:
            _draw_texts(text_values=[date_min, date_max], draw_value_4=draw_value_4, mark=1, y_low=y_low, y_up=y_up,
                        date_flag=True)
        else:
            _draw_texts(text_values=[value_min, value_mean, value_median, value_max], draw_value_4=draw_value_4,
                        mark=1, y_low=y_low, y_up=y_up)
    plt.show()


def _check_numeric(col, _value_df, img_dir, date_flag=False):
    """
    Summarize numeric feature

    Parameters
    ----------
    col: name of the column
    _value_df: pandas dataframe
        the dataframe that contains the values of 'col'
    img_dir: root directory for the generated images
    date_flag: boolean
        Whether the column is date type

    Returns
    -------
    Dictionary: (column value, result dataframe / error message)
    """

    value_df = _value_df.copy()

    # ensure all values are numeric
    value_df[col] = pd.to_numeric(value_df[col], errors='coerce')

    # percentage of nan
    nan_rate = value_df[value_df[col].isnull()].shape[0] * 1.0 / value_df.shape[0]

    # if all values are nan
    if nan_rate == 1:
        return {'column': col, 'error_msg': 'all values are nan'}
    else:
        # unique_level
        num_uni = value_df[col].dropna().nunique()

        # get clean values
        value_dropna = value_df[col].dropna().values

        # get sample value
        sample_value = np.random.choice(value_dropna, 1)[0]

        # get min, mean, median, max
        value_min = value_df[col].min()
        value_mean = value_df[col].mean()
        value_median = value_df[col].median()
        value_max = value_df[col].max()
        if date_flag:
            date_min = pd.to_datetime(value_df[col.replace('_numeric', '')], errors='coerce').min()
            date_max = pd.to_datetime(value_df[col.replace('_numeric', '')], errors='coerce').max()

        # get distribution
        scale_flg = 0
        draw_values = value_dropna
        draw_value_4 = [value_min, value_mean, value_median, value_max]
        if np.max([abs(value_min), abs(value_max)]) >= pow(10, 6):
            scale_flg = 1
            draw_values, draw_value_4 = _get_scale_draw_values(draw_values, draw_value_4)

        # draw and save distribution graph
        dpi = 72
        if date_flag:
            plt.figure(figsize=(635. / dpi, 635. / (9. / 5.5) / dpi), dpi=dpi)
        else:
            plt.figure(figsize=(635. / dpi, 635. / (9. / 4.5) / dpi), dpi=dpi)
        if scale_flg:
            plt.title('%s (log10 scale)' %(col))
        else:
            plt.title('%s' %(col))

        # if unique level is less than 10, draw countplot instead
        if num_uni <= 10:
            temp_df = pd.DataFrame(draw_values, columns=['value'])
            sns.countplot(temp_df['value'], color=DIS_LINE)
            if num_uni > 5:
                plt.xticks(rotation=90)
        else:
            ax = sns.distplot(draw_values, color=DIS_LINE, norm_hist=True, hist=False)
            y_low, y_up = ax.get_ylim()

            if date_flag:
                _draw_texts(text_values=[date_min, date_max], draw_value_4=draw_value_4, mark=1, y_low=y_low, y_up=y_up,
                            date_flag=True)
            else:
                _draw_texts(text_values=[value_min, value_mean, value_median, value_max], draw_value_4=draw_value_4,
                            mark=1, y_low=y_low, y_up=y_up)

        # save the graphs
        # adjust graph name
        graph_name = col
        if '/' in graph_name:
            graph_name = graph_name.replace('/', '')
        plt.savefig(os.path.join(img_dir, graph_name + '.png'), transparent=True, dpi=dpi)

        output = [
            {'feature': 'column', 'value': col, 'graph': 'Distribution'},
            {'feature': 'sample_value', 'value': sample_value},
            {'feature': 'nan_rate', 'value': nan_rate},
            {'feature': 'num_uni', 'value': '%d/%d' %(num_uni, len(value_dropna))},
            {'feature': 'value_min', 'value': value_min},
            {'feature': 'value_mean', 'value': value_mean},
            {'feature': 'value_median', 'value': value_median},
            {'feature': 'value_max', 'value': value_max},
        ]

        if date_flag:
            output.append({'feature': 'date_min', 'value': date_min})
            output.append({'feature': 'date_max', 'value': date_max})

        return {'column': col, 'result_df': pd.DataFrame(output)}


def _check_string(col, _value_df):
    """
    Summarize string feature

    Parameters
    ----------
    col: name of the column
    _value_df: pandas dataframe
        the dataframe that contains the values of 'col'

    Returns
    -------
    Dictionary: (column value, [result dataframe, top 10 value counts dataframe] / error message)
    """

    value_df = _value_df.copy()

    # percentage of nan
    nan_rate = value_df[value_df[col].isnull()].shape[0] * 1.0 / value_df.shape[0]

    # if all sample values are nan
    if nan_rate == 1:
        return {'column': col, 'error_msg': 'all values are nan'}
    else:
        # unique_level
        num_uni = value_df[col].dropna().nunique()

        # get clean values
        value_dropna = value_df[col].dropna().values

        # get sample value
        sample_value = np.random.choice(value_dropna, 1)[0]

        # get the top 10 value counts
        value_counts_df = pd.DataFrame(pd.Series(value_df[col].values).value_counts(dropna=False), columns=['count'])
        value_counts_df[col] = value_counts_df.index.values
        value_counts_df[col] = value_counts_df[col].map(str)
        value_counts_df = value_counts_df.sort_values(by='count', ascending=False).head(10)[[col, 'count']]

        output = [
            {'feature': 'column', 'value': col},
            {'feature': 'sample_value', 'value': sample_value},
            {'feature': 'nan_rate', 'value': nan_rate},
            {'feature': 'num_uni', 'value': '%d/%d' %(num_uni, len(value_dropna))}
        ]
        return {'column': col, 'result_df': [pd.DataFrame(output), value_counts_df]}


def _check_date(col, value_df, img_dir):
    """
    Summarize date feature

    Parameters
    ----------
    col: name of the column
    _value_df: pandas dataframe
        the dataframe that contains the values of 'col'
    img_dir: root directory for the generated images

    Returns
    -------
    Dictionary: (column value, [result dataframe, top 10 value counts dataframe] / error message)
    """

    numeric_output = _check_numeric(col, value_df, img_dir, date_flag=True)
    col = numeric_output['column']
    if 'result_df' in numeric_output.keys():
        result_df = numeric_output['result_df']
        result_df.loc[result_df['feature']=='column', 'value'] = col.replace('_numeric', '')
        result_df.loc[0, 'graph'] = 'Distribution (months)'

        return {'column': col.replace('_numeric', ''), 'result_df': result_df}
    else:
        return {'column': col.replace('_numeric', ''), 'error_msg': numeric_output['error_msg']}


def _insert_string_results(string_results, ws, row_height):
    """
    Insert result of a string type column into a worksheet

    Parameters
    ----------
    string_results: dict
        The result dictionary
    ws: Excel worksheet instance
    row_height: float
        Height of the rows
    """

    # construct thin border
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # loop and output result
    for result in string_results:
        column = result['column']
        if 'result_df' not in result.keys():
            ws.append([column, result['error_msg']])
            for col in ['A', 'B']:
                ws['%s%d' %(col, ws.max_row)].style = 'Bad'
            ws.append([''])
            continue
        result_df = result['result_df'][0]
        value_counts_df = result['result_df'][1]
        head_row = _insert_df(result_df, ws)

        # if there is value counts result
        if len(value_counts_df) > 0:
            value_counts_df = value_counts_df.rename(columns={column: 'top 10 values'})
            databar_head = _insert_df(value_counts_df, ws, header=True, head_style='60 % - Accent5')

            # add conditional formatting: data bar
            first = FormatObject(type='num', val=0)
            second = FormatObject(type='num', val=value_counts_df['count'].max())
            data_bar = DataBar(cfvo=[first, second], color=DIS_LINE.replace('#', ''),
                               showValue=True, minLength=None, maxLength=None)

            # assign the data bar to a rule
            rule = Rule(type='dataBar', dataBar=data_bar)
            ws.conditional_formatting.add('B%d:B%d' %(databar_head+1, databar_head+len(value_counts_df)), rule)

            # draw the thick outline border
            _style_range(ws, 'A%d:B%d'%(head_row, databar_head+len(value_counts_df)), border=border)
        else:
            _style_range(ws, 'A%d:B%d'%(head_row, head_row+result_df.shape[0]-1), border=border)

        # add gap
        ws.append([''])

    # adjust the worksheet
    _adjust_ws(ws=ws, row_height=row_height)


def _check_features(table_schema):
    """
    Get list of features to check based on table schema

    Parameters
    ----------
    table_schema: table schema

    Returns
    -------
    exclude_features:
        list of features to exclude during the checking process
    check_features:
        a dictionary of (data type, feature list) to check
    """
    # exclude some features where include equals to 0
    exclude_features = list(table_schema[table_schema['include'] == 0]['column'].values)

    # classify features based on data type
    table_schema_include = table_schema[table_schema['include'] == 1].reset_index(drop=True)
    correct_types = ['key', 'numeric', 'str', 'date']
    check_features = {}
    for correct_type in correct_types:
        check_features[correct_type] = table_schema_include[table_schema_include['type'] == correct_type]['column'].values

    return exclude_features, check_features


def data_summary(table_schema, table, fname, sample_size=1.0, sample_rows=100, output_root='', keep_images=False, n_jobs=1):
    """
    Summarize basic information of all columns in a data table
    based on the provided data schema

    Parameters
    ----------
    table_schema: pandas DataFrame
        schema of the table, should contain data types of each column
    table: pandas DataFrame
        the data table
    fname: string
        the output file name
    sample_size: integer or float(<=1.0), default=1.0
        int: number of sample rows to do the summary (useful for large tables)
        float: sample size in percentage
    sample_rows: integer
        number of rows to get data samples
    output_root: string
        the root directory for the output file
    keep_images: boolean
        whether to keep all generated images
    n_jobs: int
        the number of jobs to run in parall
    """

    # check sample_size
    if sample_size > 1:
        if int(sample_size) != sample_size:
            raise ValueError('sample_size: only accept integer when it is > 1.0')
        if sample_size > table.shape[0]:
            print("sample_size: %d is larger than the data size: %d" % (sample_size, table.shape[0]))

    # check output_root
    if output_root != '':
        if not os.path.isdir(output_root):
            raise ValueError('output_root: root not exists')

    # get data samples before sample_size
    data_sample = table.sample(sample_rows).reset_index(drop=True)

    # calculate the sample size
    if sample_size <= 1.0:
        sample_size = int(table.shape[0] * sample_size)
    if sample_size < table.shape[0]:
        table = table.sample(sample_size).reset_index(drop=True)

    exclude_features, check_features = _check_features(table_schema)

    # temp dir to store all the images generated
    img_dir = 'img_temp'
    if os.path.isdir(img_dir):
        shutil.rmtree(img_dir)
    os.mkdir(img_dir)

    # create a new workbook to store everything
    wb = openpyxl.Workbook()

    # list of results
    all_results = []

    # key features
    key_features = check_features['key']
    if len(key_features) > 0:
        # get the check result
        _n_jobs = np.min([n_jobs, len(key_features)])
        key_results = Parallel(n_jobs=_n_jobs)(delayed(_check_string)(col, table[[col]]) for col in key_features)
        all_results += key_results

        ws = wb.create_sheet(title=u'key')
        # write the final result to work sheet
        _insert_string_results(key_results, ws, 25)

    # numeric features
    numeric_features = check_features['numeric']
    if len(numeric_features) > 0:
        # get the check result
        _n_jobs = np.min([n_jobs, len(numeric_features)])
        numeric_results = Parallel(n_jobs=_n_jobs)(
            delayed(_check_numeric)(col, table[[col]], img_dir) for col in numeric_features)
        all_results += numeric_results

        ws = wb.create_sheet(title=u'numeric')
        # write the final result to work sheet
        _insert_numeric_results(numeric_results, ws, 35, img_dir)

    # string features
    string_features = check_features['str']
    if len(string_features) > 0:
        _n_jobs = np.min([n_jobs, len(string_features)])
        string_results = Parallel(n_jobs=_n_jobs)(delayed(_check_string)(col, table[[col]]) for col in string_features)
        all_results += string_results

        ws = wb.create_sheet(title=u'string')
        # write the final result to work sheet
        _insert_string_results(string_results, ws, 25)

    # date features
    date_features = check_features['date']
    if len(date_features) > 0:
        # get the current time
        snapshot_date_now = str(datetime.datetime.now().date())
        for col in date_features:
            table['%s_numeric' %(col)] = (pd.to_datetime(snapshot_date_now) - pd.to_datetime(table[col], errors='coerce')).astype('timedelta64[M]', errors='ignore')
        _n_jobs = np.min([n_jobs, len(date_features)])
        date_results = Parallel(n_jobs=_n_jobs)(
            delayed(_check_date)('%s_numeric' % col, table[['%s_numeric' % col, col]], img_dir) for col in date_features)
        all_results += date_results

        ws = wb.create_sheet(title=u'date')
        # write the final result to work sheet
        _insert_numeric_results(date_results, ws, 35, img_dir, date_flag=True)

    # write schema
    ws = wb['Sheet']
    ws.title = 'summary'
    out_schema = table_schema[['column', 'type']]
    out_schema['check'] = 'Ok'

    # output error features
    error_indices = []
    if len(exclude_features) > 0:
        out_schema['check'] = out_schema.apply(lambda x : 'exclude' if x['column'] in exclude_features else x['check'], axis=1)
        error_indices += list(out_schema[out_schema['column'].isin(exclude_features)].index.values)

    # tidy up the output
    error_msg_dict = {}
    correct_info = []
    for result in all_results:
        if 'error_msg' in result.keys():
            error_msg_dict[result['column']] = result['error_msg']
        else:
            if type(result['result_df']) == list:
                result_df = result['result_df'][0]
            else:
                result_df = result['result_df']
            info = pd.Series(result_df['value'].values, index=result_df['feature']).to_dict()
            correct_info.append(info)

    correct_info_df = pd.DataFrame(correct_info)
    out_schema = out_schema.merge(correct_info_df, on='column', how='left')

    if len(error_msg_dict) > 0:
        out_schema['check'] = out_schema.apply(lambda x : error_msg_dict[x['column']] if x['column'] in error_msg_dict.keys() else x['check'], axis=1)
        error_indices += list(out_schema[out_schema['column'].isin(error_msg_dict.keys())].index.values)

    if 'date_min' in out_schema.columns.values:
        order_columns = ['column', 'type', 'check', 'sample_value', 'nan_rate', 'num_uni', 'value_min', 'value_mean',
                         'value_median', 'value_max', 'date_min', 'date_max']
    else:
        order_columns = ['column', 'type', 'check', 'sample_value', 'nan_rate', 'num_uni', 'value_min', 'value_mean',
                         'value_median', 'value_max']

    _ = _insert_df(out_schema[order_columns], ws, header=True)
    if len(error_indices) > 0:
        for idx in error_indices:
            ws['C%d' %(idx+2)].style = 'Bad'

    _adjust_ws(ws=ws, row_height=25)

    # write data samples
    ws = wb.create_sheet(title=u'sample')
    _ = _insert_df(data_sample, ws, header=True, head_color=True, bold_first_column=False)
    _adjust_ws(ws=ws, row_height=20)

    wb.save(filename=os.path.join(output_root, 'data_summary_%s.xlsx' %(fname)))

    # remove all temp images
    if not keep_images:
        shutil.rmtree(img_dir)


def data_summary_notebook(table_schema, table, fname, output_root=''):
    """
    Automatically generate ipynb for data summary process

    Parameters
    ----------
    table_schema: pandas DataFrame
        schema of the table, should contain data types of each column
    table: pandas DataFrame
        the data table
    fname: string
        the output file name
    output_root: string, default=''
        the root directory for the output file
    """

    # check output_root
    if output_root != '':
        if not os.path.isdir(output_root):
            raise ValueError('output_root: root not exists')

    # generate output file path
    output_path = os.path.join(output_root, 'data_summary_notebook_%s.py' %(fname))

    # delete potential generated script and notebook
    if os.path.isfile(output_path):
        os.remove(output_path)

    if os.path.isfile(output_path.replace('.py', '.ipynb')):
        os.remove(output_path.replace('.py', '.ipynb'))

    dir_path = os.path.dirname(os.path.realpath(__file__))
    main_line = open(dir_path + '/templates/data_summary_main.txt').read()
    key_str_line = open(dir_path + '/templates/data_summary_key_str.txt').read()
    numeric_line = open(dir_path + '/templates/data_summary_numeric.txt').read()
    date_line = open(dir_path + '/templates/data_summary_date.txt').read()

    exclude_features, _ = _check_features(table_schema)

    with open(output_path, "a") as outbook:
        # main
        outbook.write(main_line)

        # output exclude features
        if len(exclude_features) > 0:
            outbook.write('\n"""\n## excluded columns\n\n')
            outbook.write('%s\n"""\n\n' % (', '.join(exclude_features)))

        # columns follow the order from table
        check_cols = [col for col in table.columns.values if col not in exclude_features]
        for col in check_cols:
            # get the data type of the column
            col_type = table_schema[table_schema['column'] == col]['type'].values[0]
            outbook.write('\n"""\n## %s (type: %s)\n\n"""\n\n' %(col, col_type))
            outbook.write('col = "%s"\n' %(col))

            # for key and str, check simple value counts
            if (col_type == 'key') or (col_type == 'str'):
                outbook.write(key_str_line)
            elif col_type == 'date':
                outbook.write(date_line)
            else:
                outbook.write(numeric_line)

        outbook.close()

    os.system("python -m py2nb %s %s" %(output_path, output_path.replace('.py', '.ipynb')))

