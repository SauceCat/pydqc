import pandas as pd
import numpy as np
import os
import shutil

import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.formatting.rule import DataBar, FormatObject, Rule

from sklearn.externals.joblib import Parallel, delayed
from scipy.stats import spearmanr

import matplotlib.pyplot as plt
import seaborn as sns 
sns.set_style('white')

from matplotlib_venn import venn2
import datetime

from .dqc_utils import (
    _style_range, _get_scale_draw_values, _draw_texts,
    _adjust_ws, _insert_df, _insert_numeric_results
)

import warnings
warnings.filterwarnings('ignore')


# global color values
TABLE1_DARK = "#4BACC6"
TABLE1_LIGHT = "#DAEEF3"

TABLE2_DARK = "#F79646"
TABLE2_LIGHT = "#FDE9D9"


def distribution_compare_pretty(_df1, _df2, col, figsize=None, date_flag=False):
    """
    Draw pretty distribution graph for data compare

    Parameters
    ----------
    _df1: pandas DataFrame
        slice of table1 containing enough information to check
    _df2: pandas DataFrame
        slice of table2 containing enough information to check
    col: string
        name of column to check
    figsize: tuple, default=None
        figure size
    date_flag: bool, default=False
        whether it is checking date features
    """

    # color values for graph
    TABLE1_DARK = "#4BACC6"
    TABLE2_DARK = "#F79646"

    df1, df2 = _df1.copy(), _df2.copy()

    if date_flag:
        numeric_col = '%s_numeric' %(col)
        if numeric_col not in df1.columns.values:
            snapshot_date_now = str(datetime.datetime.now().date())
            df1[numeric_col] = (pd.to_datetime(snapshot_date_now) - pd.to_datetime(df1[col],
                errors='coerce')).astype('timedelta64[M]', errors='ignore')
        if numeric_col not in df2.columns.values:
            snapshot_date_now = str(datetime.datetime.now().date())
            df2[numeric_col] = (pd.to_datetime(snapshot_date_now) - pd.to_datetime(df2[col],
                errors='coerce')).astype('timedelta64[M]', errors='ignore')
    else:
        numeric_col = col

    value_mins = [df1[numeric_col].min(), df2[numeric_col].min()]
    value_means = [df1[numeric_col].mean(), df2[numeric_col].mean()]
    value_medians = [df1[numeric_col].median(), df2[numeric_col].median()]
    value_maxs = [df1[numeric_col].max(), df2[numeric_col].max()]

    if date_flag:
        date_mins = [pd.to_datetime(df1[col], errors='coerce').min(), pd.to_datetime(df2[col], errors='coerce').min()]
        date_maxs = [pd.to_datetime(df1[col], errors='coerce').max(), pd.to_datetime(df2[col], errors='coerce').max()]

    both_value_max = np.max([abs(v) for v in value_maxs] + [abs(v) for v in value_mins])

    # get clean values
    df1_sample_dropna_values = df1[numeric_col].dropna().values
    df2_sample_dropna_values = df2[numeric_col].dropna().values

    # get distribution
    scale_flg = 0
    df1_draw_values = df1_sample_dropna_values
    df1_draw_value_4 = [value_mins[0], value_means[0], value_medians[0], value_maxs[0]]

    df2_draw_values = df2_sample_dropna_values
    df2_draw_value_4 = [value_mins[1], value_means[1], value_medians[1], value_maxs[1]]

    if both_value_max >= pow(10, 6):
        scale_flg = 1
        df1_draw_values, df1_draw_value_4 = _get_scale_draw_values(df1_draw_values, df1_draw_value_4)
        df2_draw_values, df2_draw_value_4 = _get_scale_draw_values(df2_draw_values, df2_draw_value_4)

    # draw the graph
    plt.clf()
    if figsize is not None:
        plt.figure(figsize)
    else:
        plt.figure(figsize=(10, 5))

    if scale_flg:
        plt.title('%s (log10 scale)' %(col))
    else:
        plt.title('%s' %(col))

    # if unique level is less than 10, draw countplot instead
    both_num_uni = np.max([df1[col].dropna().nunique(), df2[col].dropna().nunique()])
    if both_num_uni <= 10:
        df1_temp = pd.DataFrame(df1_sample_dropna_values, columns=['value'])
        df1_temp['type'] = 'table1'
        df2_temp = pd.DataFrame(df2_sample_dropna_values, columns=['value'])
        df2_temp['type'] = 'table2'
        full_temp = pd.concat([df1_temp, df2_temp], axis=0)
        sns.countplot(full_temp['value'], hue=full_temp['type'], palette=sns.color_palette([TABLE1_DARK, TABLE2_DARK]))
        if both_num_uni > 5:
            plt.xticks(rotation=90)
        plt.legend(loc=1)
    else:
        ax1 = sns.distplot(df1_draw_values, color=TABLE1_DARK, hist=False, label='table1')
        ax2 = sns.distplot(df2_draw_values, color=TABLE2_DARK, hist=False, label='table2')
        y_low_1, y_up_1 = ax1.get_ylim()
        y_low_2, y_up_2 = ax2.get_ylim()
        y_low, y_up = np.min([y_low_1, y_low_2]), np.max([y_up_1, y_up_2])
        plt.ylim((y_low, y_up))

        if date_flag:
            _draw_texts(text_values=[date_mins[0], date_maxs[0]], draw_value_4=df1_draw_value_4, mark=1,
                        y_low=y_low, y_up=y_up, date_flag=True)
            _draw_texts(text_values=[date_mins[1], date_maxs[1]], draw_value_4=df2_draw_value_4, mark=2,
                        y_low=y_low, y_up=y_up, date_flag=True)
        else:
            _draw_texts(text_values=[value_mins[0], value_means[0], value_medians[0], value_maxs[0]],
                        draw_value_4=df1_draw_value_4, mark=1, y_low=y_low, y_up=y_up)
            _draw_texts(text_values=[value_mins[1], value_means[1], value_medians[1], value_maxs[1]],
                        draw_value_4=df2_draw_value_4, mark=2, y_low=y_low, y_up=y_up)

    plt.show()


def _simple_stats(col, _df1, _df2, stat_type):
    """
    Check simple statistical information

    Parameters
    ----------
    col: string
        name of column to check
    _df1: pandas DataFrame
        slice of table1 containing enough information to check
    _df2: pandas DataFrame
        slice of table2 containing enough information to check
    stat_type: type of the column

    Returns
    -------
    output: dictionary contains the output result
    """

    df1 = _df1.copy()
    df2 = _df2.copy()

    # default output
    output = {'sample_value': np.nan, 'nan_rate': np.nan, 'num_uni': np.nan,
              'value_min': np.nan, 'value_mean': np.nan, 'value_median': np.nan, 'value_max': np.nan,
              'date_min': np.nan, 'date_max': np.nan}

    # nan_rate
    nan_rate1 = df1[df1[col].isnull()].shape[0] * 1.0 / df1.shape[0]
    nan_rate2 = df2[df2[col].isnull()].shape[0] * 1.0 / df2.shape[0]
    output['nan_rate'] = [nan_rate1, nan_rate2]

    if nan_rate1 == 1 or nan_rate2 == 1:
        return output

    # sample value
    output['sample_value'] = [df1[col].dropna().sample(1).values[0], df2[col].dropna().sample(1).values[0]]

    # num_uni
    output['num_uni'] = [df1[col].dropna().nunique(), df2[col].dropna().nunique()]

    if (stat_type == 'key') or (stat_type == 'str'):
        return output

    # stats
    output['value_min'] = [df1[col].min(), df2[col].min()]
    output['value_mean'] = [df1[col].mean(), df2[col].mean()]
    output['value_median'] = [df1[col].median(), df2[col].median()]
    output['value_max'] = [df1[col].max(), df2[col].max()]

    if stat_type == 'numeric':
        return output

    # date_min
    date_min1 = pd.to_datetime(df1[col.replace('_numeric', '')], errors='coerce').min()
    date_min2 = pd.to_datetime(df2[col.replace('_numeric', '')], errors='coerce').min()
    output['date_min'] = [date_min1, date_min2]

    # date_max
    date_max1 = pd.to_datetime(df1[col.replace('_numeric', '')], errors='coerce').max()
    date_max2 = pd.to_datetime(df2[col.replace('_numeric', '')], errors='coerce').max()
    output['date_max'] = [date_max1, date_max2]

    return output


def _compare_key(key, _df1, _df2, img_dir):
    """
    Compare two key type values

    Parameters
    ----------
    key: string
        name of column to check
    _df1: pandas DataFrame
        slice of table1 containing enough information to check
    _df2: pandas DataFrame
        slice of table2 containing enough information to check
    img_dir: root directory for the generated images

    Returns
    -------
    Dictionary contains the output result
    """

    df1 = _df1.copy()
    df2 = _df2.copy()

    # get basic stats information
    stat_output = _simple_stats(key, df1, df2, 'key')

    # basic check for key
    nan_rate1, nan_rate2 = stat_output['nan_rate']

    if (nan_rate1 == 1) or (nan_rate2 == 1):
        if (nan_rate1 == 1) and (nan_rate2 == 1):
            error_msg = 'all nan in both table'
        elif nan_rate1 == 1:
            error_msg = 'all nan in table1'
        else:
            error_msg = 'all nan in table2'
        return {'column': key, 'error_msg': error_msg}

    set_df1_key = set(df1[key].dropna().values) if nan_rate1 < 1 else set()
    set_df2_key = set(df2[key].dropna().values) if nan_rate2 < 1 else set()
    key_overlap = len(set_df1_key.intersection(set_df2_key))
    key_only_df1, key_only_df2 = len(set_df1_key - set_df2_key), len(set_df2_key - set_df1_key)
    overlap_rate = key_overlap * 1.0 / (key_overlap + key_only_df1 + key_only_df2)

    # generate the output
    output = [
        {'feature': 'column', 'value': key, 'graph': 'venn graph'},
        {'feature': 'sample_value', 'value': '\n'.join([str(v) for v in stat_output['sample_value']])},
        {'feature': 'nan_rate', 'value': '\n'.join([str(round(v, 3)) for v in stat_output['nan_rate']])},
        {'feature': 'num_uni', 'value': '%s/%s\n%s/%s' % (str(stat_output['num_uni'][0]), str(df1.dropna().shape[0]),
                                                          str(stat_output['num_uni'][1]), str(df2.dropna().shape[0]))},
        {'feature': 'overlap', 'value': key_overlap},
        {'feature': 'only in table1', 'value': key_only_df1},
        {'feature': 'only in table2', 'value': key_only_df2},
        {'feature': 'overlap rate', 'value': round(overlap_rate, 3)}
    ]

    # draw the venn graph
    dpi = 72
    plt.figure(figsize=(635. / dpi, 635. / (9. / 5.) / dpi), dpi=dpi)
    venn2([set_df1_key, set_df2_key], set_labels=['table1', 'table2'], set_colors=(TABLE1_DARK, TABLE2_DARK), alpha=0.8)

    # save the graphs
    # adjust graph name
    graph_name = key
    if '/' in graph_name:
        graph_name = graph_name.replace('/', '')
    plt.savefig(os.path.join(img_dir, graph_name + '.png'), transparent=True, dpi=dpi)

    return {'column': key, 'result_df': pd.DataFrame(output), 'corr': {'column': key, 'corr': round(overlap_rate, 3)}}


def _compare_numeric(col, _df1, _df2, img_dir, date_flag=False):
    """
    Compare two numeric type values

    Parameters
    ----------
    col: string
        name of column to check
    _df1: pandas DataFrame
        slice of table1 containing enough information to check
    _df2: pandas DataFrame
        slice of table2 containing enough information to check
    img_dir: root directory for the generated images
    date_flag: boolean
        Whether the column is date type

    Returns
    -------
    Dictionary contains the output result
    """

    # sampling
    df1_sample = _df1.copy()
    df2_sample = _df2.copy()

    stat_output = _simple_stats(col, df1_sample, df2_sample, 'numeric')

    nan_rate1, nan_rate2 = stat_output['nan_rate']
    if (nan_rate1 == 1) or (nan_rate2 == 1):
        if (nan_rate1 == 1) and (nan_rate2 == 1):
            error_msg = 'all nan in both table'
        elif nan_rate1 == 1:
            error_msg = 'all nan in table1'
        else:
            error_msg = 'all nan in table2'
        return {'column': col, 'error_msg': error_msg}

    # generate the output
    output = [
        {'feature': 'column', 'value': col, 'graph': 'Distribution'},
        {'feature': 'sample_value', 'value': '\n'.join([str(v) for v in stat_output['sample_value']])},
        {'feature': 'nan_rate', 'value': '\n'.join([str(round(v, 3)) for v in stat_output['nan_rate']])},
        {'feature': 'num_uni', 'value': '%s/%s\n%s/%s' % (str(stat_output['num_uni'][0]), str(df1_sample.dropna().shape[0]),
                                                          str(stat_output['num_uni'][1]), str(df2_sample.dropna().shape[0]))},
        {'feature': 'value_min', 'value': '\n'.join([str(round(v, 3)) for v in stat_output['value_min']])},
        {'feature': 'value_mean', 'value': '\n'.join([str(round(v, 3)) for v in stat_output['value_mean']])},
        {'feature': 'value_median', 'value': '\n'.join([str(round(v, 3)) for v in stat_output['value_median']])},
        {'feature': 'value_max', 'value': '\n'.join([str(round(v, 3)) for v in stat_output['value_max']])}
    ]

    both_value_max = np.max([abs(v) for v in stat_output['value_max']] + [abs(v) for v in stat_output['value_min']])

    # get clean values
    df1_sample_dropna_values = df1_sample[col].dropna().values
    df2_sample_dropna_values = df2_sample[col].dropna().values

    if date_flag:
        dt1 = pd.to_datetime(df1_sample[col.replace('_numeric', '')], errors='coerce')
        dt2 = pd.to_datetime(df2_sample[col.replace('_numeric', '')], errors='coerce')
        date_min1, date_max1 = dt1.min(), dt1.max()
        date_min2, date_max2 = dt2.min(), dt2.max()

    # get distribution
    scale_flg = 0
    df1_draw_values = df1_sample_dropna_values
    df1_draw_value_4 = [stat_output['value_min'][0], stat_output['value_mean'][0],
                        stat_output['value_median'][0], stat_output['value_max'][0]]

    df2_draw_values = df2_sample_dropna_values
    df2_draw_value_4 = [stat_output['value_min'][1], stat_output['value_mean'][1],
                        stat_output['value_median'][1], stat_output['value_max'][1]]

    if both_value_max >= pow(10, 6):
        scale_flg = 1
        df1_draw_values, df1_draw_value_4 = _get_scale_draw_values(df1_draw_values, df1_draw_value_4)
        df2_draw_values, df2_draw_value_4 = _get_scale_draw_values(df2_draw_values, df2_draw_value_4)

    # calculate correlation between two distributions
    if np.max(stat_output['num_uni']) <= 100:
        vc1, vc2 = _value_counts_df(df1_draw_values), _value_counts_df(df2_draw_values)
        vc = vc1.merge(vc2, on='value', how='outer').fillna(0)
        obs1, obs2 = vc['count_x'].values * 1.0 / vc['count_x'].sum(), vc['count_y'].values * 1.0 / vc['count_y'].sum()
    else:
        both_min = np.min([np.min(df1_draw_values), np.min(df2_draw_values)])
        both_max = np.max([np.max(df1_draw_values), np.max(df2_draw_values)])
        hist1 = np.histogram(df1_draw_values, bins=100, range=(both_min, both_max), normed=False, density=False)
        hist2 = np.histogram(df2_draw_values, bins=100, range=(both_min, both_max), normed=False, density=False)
        obs1, obs2 = hist1[0] / (np.sum(hist1[0]) * 1.0), hist2[0] / (np.sum(hist2[0]) * 1.0)

    if len(obs1) == 1:
        corr = np.min([1. - nan_rate1, 1. - nan_rate2]) * 1.0 / np.max([1. - nan_rate1, 1. - nan_rate2])
    elif list(obs1) == list(obs2):
        corr = 1.0
    else:
        corr = spearmanr(obs1, obs2)[0]

    # draw and save distribution graph
    dpi = 72
    if date_flag:
        plt.figure(figsize=(635. / dpi, 635. / (9. / 8.) / dpi), dpi=dpi)
    else:
        plt.figure(figsize=(635. / dpi, 635. / (9. / 6.) / dpi), dpi=dpi)
    if scale_flg:
        plt.title('%s (log10 scale)' %(col))
    else:
        plt.title('%s' %(col))

    # if unique level is less than 10, draw countplot instead
    both_num_uni = np.max(stat_output['num_uni'])
    if both_num_uni <= 10:
        df1_temp = pd.DataFrame(df1_sample_dropna_values, columns=['value'])
        df1_temp['type'] = 'table1'
        df2_temp = pd.DataFrame(df2_sample_dropna_values, columns=['value'])
        df2_temp['type'] = 'table2'
        full_temp = pd.concat([df1_temp, df2_temp], axis=0)
        sns.countplot(full_temp['value'], hue=full_temp['type'], palette=sns.color_palette([TABLE1_DARK, TABLE2_DARK]))
        if both_num_uni > 5:
            plt.xticks(rotation=90)
        plt.legend(loc=1)
    else:
        ax1 = sns.distplot(df1_draw_values, color=TABLE1_DARK, hist=False, label='table1')
        ax2 = sns.distplot(df2_draw_values, color=TABLE2_DARK, hist=False, label='table2')
        y_low_1, y_up_1 = ax1.get_ylim()
        y_low_2, y_up_2 = ax2.get_ylim()
        y_low, y_up = np.min([y_low_1, y_low_2]), np.max([y_up_1, y_up_2])
        plt.ylim((y_low, y_up))

        if date_flag:
            _draw_texts(text_values=[date_min1, date_max1], draw_value_4=df1_draw_value_4, mark=1,
                        y_low=y_low, y_up=y_up, date_flag=True)
            _draw_texts(text_values=[date_min2, date_max2], draw_value_4=df2_draw_value_4, mark=2,
                        y_low=y_low, y_up=y_up, date_flag=True)
        else:
            _draw_texts(text_values=[stat_output['value_min'][0], stat_output['value_mean'][0],
                                     stat_output['value_median'][0], stat_output['value_max'][0]],
                        draw_value_4=df1_draw_value_4, mark=1, y_low=y_low, y_up=y_up)
            _draw_texts(text_values=[stat_output['value_min'][1], stat_output['value_mean'][1],
                                     stat_output['value_median'][1], stat_output['value_max'][1]],
                        draw_value_4=df2_draw_value_4, mark=2, y_low=y_low, y_up=y_up)

    # save the graphs
    # adjust graph name
    graph_name = col
    if '/' in graph_name:
        graph_name = graph_name.replace('/', '')
    plt.savefig(os.path.join(img_dir, graph_name + '.png'), transparent=True, dpi=dpi)

    if date_flag:
        output.append({'feature': 'date_min', 'value': '%s\n%s' %(date_min1, date_min2)})
        output.append({'feature': 'date_max', 'value': '%s\n%s' %(date_max1, date_max2)})
    output.append({'feature': 'corr', 'value': round(corr, 3)})

    return {'column': col, 'result_df': pd.DataFrame(output), 'corr': {'column': col, 'corr': round(corr, 3)}}


def _value_counts_df(values):
    """
    Construct value count dataframe

    Parameters
    ----------
    values: arrary_like
        values to construct the value count

    Returns
    -------
    value counts dataframe
    """
    temp = pd.DataFrame(pd.Series(values).value_counts(dropna=False), columns=['count'])
    temp['value'] = temp.index.values
    temp['value'] = temp['value'].map(str)
    return temp.reset_index(drop=True)


def _compare_string(col, _df1, _df2):
    """
    Compare two string type values

    Parameters
    ----------
    col: string
        name of column to check
    _df1: pandas DataFrame
        slice of table1 containing enough information to check
    _df2: pandas DataFrame
        slice of table2 containing enough information to check

    Returns
    -------
    Dictionary contains the output result
    """

    # sampling
    df1_sample = _df1.copy()
    df2_sample = _df2.copy()

    # get basic stats information
    stat_output = _simple_stats(col, df1_sample, df2_sample, 'str')
    nan_rate1, nan_rate2 = stat_output['nan_rate']

    if (nan_rate1 == 1) or (nan_rate2 == 1):
        if (nan_rate1 == 1) and (nan_rate2 == 1):
            error_msg = 'all nan in both table'
        elif nan_rate1 == 1:
            error_msg = 'all nan in table1'
        else:
            error_msg = 'all nan in table2'
        return {'column': col, 'error_msg': error_msg}

    # basic check for category features
    set_df1_col = set(df1_sample[col].dropna().values) if nan_rate1 < 1 else set()
    set_df2_col = set(df2_sample[col].dropna().values) if nan_rate2 < 1 else set()
    col_overlap = len(set_df1_col.intersection(set_df2_col))
    col_only_df1, col_only_df2 = len(set_df1_col - set_df2_col), len(set_df2_col - set_df1_col)

    # generate the output
    output = [
        {'feature': 'column', 'value': col, 'graph': ''},
        {'feature': 'sample_value', 'value': '\n'.join([str(v) for v in stat_output['sample_value']])},
        {'feature': 'nan_rate', 'value': '\n'.join([str(round(v, 3)) for v in stat_output['nan_rate']])},
        {'feature': 'num_uni', 'value': '%s/%s\n%s/%s' % (str(stat_output['num_uni'][0]), str(df1_sample.dropna().shape[0]),
                                                          str(stat_output['num_uni'][1]), str(df2_sample.dropna().shape[0]))},
        {'feature': 'overlap', 'value': col_overlap},
        {'feature': 'only in table1', 'value': col_only_df1},
        {'feature': 'only in table2', 'value': col_only_df2},
    ]

    # draw the count graph
    value_counts_df1 = _value_counts_df(df1_sample[col].values)
    value_counts_df2 = _value_counts_df(df2_sample[col].values)

    value_counts_df1_top10 = value_counts_df1.sort_values('count', ascending=False).head(10)
    value_counts_df2_top10 = value_counts_df2.sort_values('count', ascending=False).head(10)

    value_counts_df = value_counts_df1.merge(value_counts_df2, on='value', how='outer').fillna(0)
    value_counts_df['count_x_per'] = value_counts_df['count_x'] * 1.0 / value_counts_df['count_x'].sum()
    value_counts_df['count_y_per'] = value_counts_df['count_y'] * 1.0 / value_counts_df['count_y'].sum()

    if len(value_counts_df) == 1:
        corr = np.min([1. - nan_rate1, 1. - nan_rate2]) * 1.0 / np.max([1. - nan_rate1, 1. - nan_rate2])
    elif list(value_counts_df['count_x_per'].values) == list(value_counts_df['count_y_per'].values):
        corr = 1.0
    else:
        corr = spearmanr(value_counts_df['count_x_per'].values, value_counts_df['count_y_per'].values)[0]
    output.append({'feature': 'corr', 'value': round(corr, 3)})

    value_counts_df_top10 = value_counts_df1_top10.merge(value_counts_df2_top10, on='value', how='outer').fillna(0)
    value_counts_df_top10 = value_counts_df_top10[['value', 'count_x', 'count_y']]

    return {'column': col, 'result_df': [pd.DataFrame(output), value_counts_df_top10],
            'corr': {'column': col, 'corr': round(corr, 3)}}


def _compare_date(col, _df1, _df2, img_dir):
    """
    Compare two date type values

    Parameters
    ----------
    col: string
        name of column to check
    _df1: pandas DataFrame
        slice of table1 containing enough information to check
    _df2: pandas DataFrame
        slice of table2 containing enough information to check
    img_dir: root directory for the generated images

    Returns
    -------
    Dictionary contains the output result
    """

    numeric_output = _compare_numeric(col, _df1, _df2, img_dir, date_flag=True)
    col = numeric_output['column']

    if 'result_df' in numeric_output.keys():
        result_df = numeric_output['result_df']
        result_df.loc[result_df['feature']=='column', 'value'] = col.replace('_numeric', '')
        result_df.loc[0, 'graph'] = 'Distribution (months)'

        return {'column': col.replace('_numeric', ''), 'result_df': result_df,
                'corr': {'column': col.replace('_numeric', ''), 'corr': numeric_output['corr']['corr']}}
    else:
        return {'column': col.replace('_numeric', ''), 'error_msg': numeric_output['error_msg']}


def _insert_compare_string_results(string_results, ws, row_height):
    """
    Insert string result into a worksheet

    Parameters
    ----------
    string_results: dict
        result to insert
    ws: Excel worksheet instance
    row_height: float
        Height of the row
    """

    # construct thick border
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    row_heights = {}

    # loop and output result
    for result in string_results:
        column = result['column']
        if not 'result_df' in result.keys():
            ws.append([column, result['error_msg']])
            for col in ['A', 'B']:
                ws['%s%d' %(col, ws.max_row)].style = 'Bad'
            ws.append([''])
            continue
        result_df = result['result_df'][0][['feature', 'value', 'graph']]
        value_counts_df = result['result_df'][1]
        head_row = _insert_df(result_df, ws)

        # if there is value counts result
        if len(value_counts_df) > 0:
            value_counts_df = value_counts_df.rename(columns={'value': 'top 10 values', 'count_x': 'count_1', 'count_y': 'count_2'})
            databar_head = _insert_df(value_counts_df, ws, header=True, head_style='60 % - Accent5')
            for row_idx in range(databar_head, databar_head+value_counts_df.shape[0]+1):
                row_heights[row_idx] = 25

            # add conditional formatting: data bar
            first = FormatObject(type='num', val=0)
            second = FormatObject(type='num', val=np.max([value_counts_df['count_1'].max(), value_counts_df['count_2'].max()]))
            data_bar1 = DataBar(cfvo=[first, second], color=TABLE1_DARK.replace('#', ''), showValue=True, minLength=None, maxLength=None)
            data_bar2 = DataBar(cfvo=[first, second], color=TABLE2_DARK.replace('#', ''), showValue=True, minLength=None, maxLength=None)

            # assign the data bar to a rule
            rule1 = Rule(type='dataBar', dataBar=data_bar1)
            ws.conditional_formatting.add('B%d:B%d' %(databar_head+1, databar_head+len(value_counts_df)), rule1)
            rule2 = Rule(type='dataBar', dataBar=data_bar2)
            ws.conditional_formatting.add('C%d:C%d' %(databar_head+1, databar_head+len(value_counts_df)), rule2)

            # draw the thick outline border
            _style_range(ws, 'A%d:C%d'%(head_row, databar_head+len(value_counts_df)), border=border)
        else:
            _style_range(ws, 'A%d:C%d'%(head_row, head_row+result_df.shape[0]-1), border=border)

        # add gap
        ws.append([''])

    _adjust_ws(ws, row_height=row_height, row_heights=row_heights, adjust_type='str')


def _check_features(schema1, schema2):
    """
    Get list of features to check based on table schemas

    Parameters
    ----------
    schema1: table1 schema
    schema2: table2 schema

    Returns
    -------
    schema:
        merged and modified schema
    check_features:
        a dictionary of (data type, feature list) to check
    """

    # select useful columns in schema
    schema1 = schema1[schema1['column'] != 'column not in current table'][['column', 'type', 'include']].rename(
        columns={'column': 'column_1', 'type': 'type_1', 'include': 'include_1'})
    schema2 = schema2[schema2['column'] != 'column not in current table'][['column', 'type', 'include']].rename(
        columns={'column': 'column_2', 'type': 'type_2', 'include': 'include_2'})

    # merge two schemas
    schema = schema1.merge(schema2, left_on='column_1', right_on='column_2', how='outer')

    # start to record the errors
    schema['error'] = ''

    schema.loc[schema['type_1'] != schema['type_2'], 'error'] = "inconsistent data types"
    schema.loc[schema['include_1'] == 0, 'error'] = "exclude"
    schema.loc[schema['include_1'] != schema['include_2'], 'error'] = "inconsistent include"
    schema.loc[schema['column_1'].isnull(), 'error'] = "column not in table1"
    schema.loc[schema['column_2'].isnull(), 'error'] = "column not in table2"

    # classify the features to compare
    schema_correct_include = schema[(schema['include_1'] == 1) & (schema['error'] == '')].reset_index(drop=True)

    check_features = {}
    for dtype in ['key', 'numeric', 'str', 'date']:
        check_features[dtype] = schema_correct_include[schema_correct_include['type_1'] == dtype]['column_1'].values

    return schema, check_features


def _insert_summary(wb, schema, corr_results):
    """
    Insert summary for the whole process

    Parameters
    ----------
    wb: the workbook to insert in
    schema: the schema to insert
    corr_results:
        list of correlation information for each features
    """

    # insert the summary
    ws = wb['Sheet']
    ws.title = 'summary'
    summary_df = schema.copy()
    corr_df = pd.DataFrame(corr_results).rename(columns={'column': 'column_1'})
    summary_df = summary_df.merge(corr_df, on='column_1', how='left')
    summary_df['corr'] = summary_df['corr'].fillna('error')
    summary_df['error_flg'] = summary_df['corr'].apply(lambda x: 1 if x == 'error' else 0)
    error_rows = summary_df[summary_df['error_flg'] == 1].index.values

    _ = _insert_df(summary_df[['column_1', 'column_2', 'type_1', 'type_2',
                               'include_1', 'include_2', 'corr', 'error']], ws, header=True, bold_first_column=False)

    for r_idx in error_rows:
        ws['G%d' % (r_idx + 2)].style = 'Bad'

    _adjust_ws(ws=ws, row_height=25)


def data_compare(table1, table2, schema1, schema2, fname, sample_size=1.0, output_root='', keep_images=False, n_jobs=1):
    """
    Compare columns between two tables

    Parameters
    ----------
    table1: pandas DataFrame
        one of the two tables to compare
    table2: pandas DataFrame
        one of the two tables to compare
    schema1: pandas DataFrame
        data schema (contains column names and corresponding data types) for _table1
    schema2: pandas DataFrame
        data schema (contains column names and corresponding data types) for _table2
    fname: string
        the output file name
    sample_size: integer or float(<=1.0), default=1.0
        int: number of sample rows to do the comparison (useful for large tables)
        float: sample size in percentage
    output_root: string, default=''
        the root directory for the output file
    keep_images: boolean, default=False
        whether to keep all generated images
    n_jobs: int, default=1
        the number of jobs to run in parallel
    """

    # check sample_size
    if sample_size > 1:
        if int(sample_size) != sample_size:
            raise ValueError('sample_size: only accept integer when it is > 1.0')
        if (sample_size > table1.shape[0]) or (sample_size > table2.shape[0]):
            print('sample_size: %d is smaller than %d or %d...' % (sample_size, table1.shape[0], table2.shape[0]))

    # check output_root
    if output_root != '':
        if not os.path.isdir(output_root):
            raise ValueError('output_root: root not exists')

    # start to compare with correct schemas
    # create a new workbook to store everything
    wb = openpyxl.Workbook()

    # prepare directory for generated images
    img_dir = 'img_temp'
    if os.path.isdir(img_dir):
        shutil.rmtree(img_dir)
    os.mkdir(img_dir)

    # calculate the sample size
    if sample_size <= 1.0:
        sample_size1 = int(table1.shape[0] * sample_size)
        sample_size2 = int(table2.shape[0] * sample_size)
        sample_size = np.min([sample_size1, sample_size2])

    schema, check_features = _check_features(schema1, schema2)
    corr_results = []

    # key features
    key_features = check_features['key']
    if len(key_features) > 0:
        _n_jobs = np.min([n_jobs, len(key_features)])
        key_results = Parallel(n_jobs=_n_jobs)(delayed(_compare_key)(col, table1[[col]], table2[[col]], img_dir)
            for col in key_features)

        for key_result in key_results:
            if 'corr' in key_result.keys():
                corr_results.append(key_result['corr'])

        # write all results to worksheet
        ws = wb.create_sheet(title=u'key')
        _insert_numeric_results(key_results, ws, 40, img_dir)

    # do sampling here
    if sample_size < table1.shape[0]:
        table1 = table1.sample(sample_size).reset_index(drop=True)
    if sample_size < table2.shape[0]:
        table2 = table2.sample(sample_size).reset_index(drop=True)

    # numeric features
    numeric_features = check_features['numeric']
    if len(numeric_features) > 0:
        _n_jobs = np.min([n_jobs, len(numeric_features)])
        numeric_results = Parallel(n_jobs=_n_jobs)(delayed(_compare_numeric)(col, table1[[col]], table2[[col]], img_dir)
            for col in numeric_features)

        for numeric_result in numeric_results:
            if 'corr' in numeric_result.keys():
                corr_results.append(numeric_result['corr'])

        # write all results to worksheet
        ws = wb.create_sheet(title=u'numeric')
        _insert_numeric_results(numeric_results, ws, 40, img_dir)

    # string features
    string_features = check_features['str']
    if len(string_features) > 0:
        _n_jobs = np.min([n_jobs, len(string_features)])
        string_results = Parallel(n_jobs=_n_jobs)(delayed(_compare_string)(col, table1[[col]], table2[[col]])
            for col in string_features)

        for string_result in string_results:
            if 'corr' in string_result.keys():
                corr_results.append(string_result['corr'])

        # write all results to worksheet
        ws = wb.create_sheet(title=u'string')
        _insert_compare_string_results(string_results, ws, 40)

    # date features
    date_features = check_features['date']
    if len(date_features) > 0:
        # get the current time
        snapshot_date_now = str(datetime.datetime.now().date())
        for col in date_features:
            table1['%s_numeric' %(col)] = (pd.to_datetime(snapshot_date_now) - pd.to_datetime(table1[col],
                errors='coerce')).astype('timedelta64[M]', errors='ignore')
            table2['%s_numeric' %(col)] = (pd.to_datetime(snapshot_date_now) - pd.to_datetime(table2[col],
                errors='coerce')).astype('timedelta64[M]', errors='ignore')
        _n_jobs = np.min([n_jobs, len(date_features)])
        date_results = Parallel(n_jobs=_n_jobs)(delayed(_compare_date)('%s_numeric' %(col),
            table1[['%s_numeric' %(col), col]], table2[['%s_numeric' %(col), col]], img_dir) for col in date_features)

        for date_result in date_results:
            if 'corr' in date_result.keys():
                corr_results.append(date_result['corr'])

        # write all results to worksheet
        ws = wb.create_sheet(title=u'date')
        _insert_numeric_results(date_results, ws, 40, img_dir, date_flag=True)

    # insert the summary
    _insert_summary(wb, schema, corr_results)

    wb.save(filename=os.path.join(output_root, 'data_compare_%s.xlsx' %(fname)))
    if not keep_images:
        shutil.rmtree(img_dir)


def data_compare_notebook(table1, table2, schema1, schema2, fname, output_root=''):
    """
    Automatically generate ipynb for data compare process

    Parameters
    ----------
    table1: pandas DataFrame
        one of the two tables to compare
    table2: pandas DataFrame
        one of the two tables to compare
    schema1: pandas DataFrame
        data schema (contains column names and corresponding data types) for _table1
    schema2: pandas DataFrame
        data schema (contains column names and corresponding data types) for _table2
    fname: string
        the output file name
    output_root: string, default=''
        the root directory for the output file
    """

    # check output_root
    if output_root != '':
        if not os.path.isdir(output_root):
            raise ValueError('output_root: root not exists')

    # generate output file path
    output_path = os.path.join(output_root, 'data_compare_notebook_%s.py' %(fname))

    # delete potential generated script and notebook
    if os.path.isfile(output_path):
        os.remove(output_path)

    if os.path.isfile(output_path.replace('.py', '.ipynb')):
        os.remove(output_path.replace('.py', '.ipynb'))

    schema, _ = _check_features(schema1, schema2)

    dir_path = os.path.dirname(os.path.realpath(__file__))
    main_line = open(dir_path + '/templates/data_compare_main.txt').read()
    key_line = open(dir_path + '/templates/data_compare_key.txt').read()
    str_line = open(dir_path + '/templates/data_compare_str.txt').read()
    numeric_line = open(dir_path + '/templates/data_compare_numeric.txt').read()
    date_line = open(dir_path + '/templates/data_compare_date.txt').read()

    with open(output_path, "a") as outbook:
        # main
        outbook.write(main_line)  # output exclude features

        schema_error = schema[schema['error'] != ''].reset_index(drop=True)
        if schema_error.shape[0] > 0:
            schema_error['error_column'] = schema_error['column_1']
            schema_error.loc[schema_error['error_column'].isnull(), 'error_column'] = schema_error['column_2']
            outbook.write('\n"""\n## error columns\n\n')
            for i in range(schema_error.shape[0]):
                get = schema_error.iloc[i]
                outbook.write('**%s:** %s<br>' % (get['error_column'], get['error']))
            outbook.write('"""\n\n')

        # only compare check columns in both table1 and table2, and follow the column order of table1
        check_cols = [col for col in table1.columns.values if col in schema[schema['error'] == '']['column_1'].values]
        for col in check_cols:
            # get the data type of the column
            col_type = schema[schema['column_1'] == col]['type_1'].values[0]

            outbook.write('\n"""\n## %s (type: %s)\n\n"""\n\n' %(col, col_type))
            outbook.write('col = "%s"\n' %(col))

            # for key and str, compare intersection
            if col_type == 'key':
                outbook.write(key_line)
            elif col_type == 'str':
                outbook.write(str_line)
            elif col_type == 'numeric':
                outbook.write(numeric_line)
            else:
                outbook.write(date_line)
        outbook.close()

    os.system("python -m py2nb %s %s" %(output_path, output_path.replace('.py', '.ipynb')))