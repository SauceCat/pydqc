import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter

import numpy as np
import matplotlib.pyplot as plt
import os

# global color values
TABLE1_DARK = "#4BACC6"
TABLE1_LIGHT = "#DAEEF3"

TABLE2_DARK = "#F79646"
TABLE2_LIGHT = "#FDE9D9"


def _style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    Parameters
    ----------
    ws: Excel worksheet instance
    cell_range: An excel range to style (e.g. A1:F20)
    border: An openpyxl Border
    fill: An openpyxl PatternFill or GradientFill
    font: An openpyxl Font object

    Note
    ----
    Originally from openpyxl documentation: http://openpyxl.readthedocs.io/en/default/styles.html
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border += top
    for cell in rows[-1]:
        cell.border += bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border += left
        r.border += right
        if fill:
            for c in row:
                c.fill = fill


def _get_scale_draw_values(sample_dropna_values, values_four):
    """
    Apply log10 scale on input values

    Parameters
    ----------
    sample_dropna_values: array_like
        Sample values without NaNs
    values_four: array_like
        Four values to treat separately

    Returns
    -------
    draw_values: ndarray
        The scaled values for sample_dropna_values
    draw_value_4: ndarray
        The scaled values for values_four
    """

    signs = np.sign(sample_dropna_values)
    draw_values = signs * np.log10(abs(sample_dropna_values) + 1)

    draw_value_4 = []
    for i in range(4):
        draw_value_4.append(np.sign(values_four[i]) * np.log10(abs(values_four[i]) + 1))

    return draw_values, draw_value_4


def _draw_texts(text_values, draw_value_4, mark, y_low, y_up, date_flag=False):
    """
    Draw indicator texts for a distribution graph

    Parameters
    ----------
    text_values: array_like
        Text values
    draw_value_4: array_like
        X axis values for texts
    mark: int
        Mark for choosing between color scheme 1 or 2
    y_low: float
        The minimum value of y axis
    y_up: float
        The maximum value of y axis
    date_flag: boolean
        Whether it is drawing for date type feature
    """

    color_dark = TABLE1_DARK if mark == 1 else TABLE2_DARK
    color_light = TABLE1_LIGHT if mark == 1 else TABLE2_LIGHT
    plt.axvline(x=draw_value_4[0], color=color_dark, linestyle='--', linewidth=1)
    plt.axvline(x=draw_value_4[3], color=color_dark, linestyle='--', linewidth=1)

    if date_flag:
        plt.text(draw_value_4[0], y_low + (y_up - y_low) * 0.1 * mark, 'max:' + str(text_values[1]),
            ha="center", va="center", bbox=dict(boxstyle="square", facecolor=color_light, edgecolor='none'))
        plt.text(draw_value_4[3], y_low + (y_up - y_low) * (0.6 + 0.1 * mark), 'min:' + str(text_values[0]),
            ha="center", va="center", bbox=dict(boxstyle="square", facecolor=color_light, edgecolor='none'))
    else:
        plt.axvline(x=draw_value_4[1], color=color_dark, linestyle='--', linewidth=1)
        plt.axvline(x=draw_value_4[2], color=color_dark, linestyle='--', linewidth=1)

        indicators = ['min', 'mean', 'median', 'max']
        for i in range(4):
            text = '%s:'%(indicators[i]) + str(round(text_values[i], 3))
            plt.text(draw_value_4[i], y_low + (y_up - y_low) * (0.2 * i + 0.1 * mark), text,
                ha="center", va="center", bbox=dict(boxstyle="square", facecolor=color_light, edgecolor='none'))


def _adjust_ws(ws, row_height, row_heights=None, adjust_type=None):
    """
    Adjust the height and width of columns for a worksheet

    Parameters
    ----------
    ws: Excel worksheet instance
    row_height: float
        Height of the columns
    row_heights: dict
        Dictionary of (row_idx, row_height) for special style case
    adjust_type: string or None
    """
    col_widths = {}
    for i, col in enumerate(ws.columns):
        col_name = xlsxwriter.utility.xl_col_to_name(i)
        col_widths[col_name] = 0
        for cell in col:
            cell.alignment = Alignment(horizontal='left', wrap_text=True)
            if cell:
                try:
                    cell_length = len(str(cell.value))
                except:
                    cell_length = len(cell.value)
                if cell_length > col_widths[col_name]:
                    col_widths[col_name] = cell_length

    for key in col_widths.keys():
        col_widths[key] *= 1.5

    for i, col in enumerate(range(ws.max_column)):
        col_name = xlsxwriter.utility.xl_col_to_name(i)
        ws.column_dimensions[col_name].width = np.min([np.max([15, col_widths[col_name]]), 80])

    if adjust_type == 'str':
        ws.column_dimensions['C'].width = col_widths['B']
        for i in range(ws.max_row + 1):
            try:
                ws.row_dimensions[i].height = row_heights[i]
            except:
                ws.row_dimensions[i].height = row_height
    else:
        for i in range(ws.max_row + 1):
            ws.row_dimensions[i].height = row_height


def _insert_df(result_df, ws, header=False, head_color=True, bold_first_column=True, head_style='Accent5'):
    """
    Insert a pandas dataframe into a worksheet

    Parameters
    ----------
    result_df: pandas Dataframe
        The dataframe
    ws: Excel worksheet instance
    header: boolean
        Whether to insert the dataframe header
    head_color: boolean
        Whether to color the dataframe header
    head_style: Excel cell style
        Cell style to apply on the header

    Returns
    -------
    head_row: int
        Row index of the header
    """

    max_col = result_df.shape[1]
    for r_idx, r in enumerate(dataframe_to_rows(result_df, index=False, header=header)):
        ws.append(r)
        for cell_idx, cell in enumerate(ws.iter_cols(max_col=max_col, min_row=ws.max_row, max_row=ws.max_row)):
            cell = cell[0]
            cell.font = Font(name='Calibri', size=11)

            # apply cell style on header
            # get the header row index
            if r_idx == 0:
                if head_color:
                    cell.style = head_style
                head_row = ws.max_row
            else:
                # bold the first column
                if cell_idx == 0 and bold_first_column:
                    cell.font = Font(bold=True)
    return head_row


def _insert_numeric_results(numeric_results, ws, row_height, img_dir, date_flag=False):
    """
    Insert results of a numeric feature into a worksheet

    Parameters
    ----------
    numeric_results: dict
        results to insert
    ws: Excel worksheet instance
    row_height: float
        Height of the row
    img_dir: string
        Root directory for the generated images
    date_flag: boolean
        Whether the feature is date type
    """

    # construct the thin border
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # loop and output the results
    for result in numeric_results:
        column = result['column']
        if 'result_df' not in result.keys():
            ws.append([column, result['error_msg']])
            for col in ['A', 'B']:
                ws['%s%d' %(col, ws.max_row)].style = 'Bad'
            ws.append([''])
            continue

        result_df = result['result_df']
        result_df = result_df[['feature', 'value', 'graph']]
        head_row = _insert_df(result_df, ws)

        # merge cells for the graph
        ws.merge_cells('C%d:C%d' %(head_row+1, head_row+result_df.shape[0]-1))
        _style_range(ws, 'A%d:C%d'%(head_row, head_row+result_df.shape[0]-1), border=border)
        ws['C%d' %(head_row+1)].border = Border(top=None, left=None, right=thin, bottom=thin)

        # add gap
        ws.append([''])

        # insert graph
        try:
            if date_flag:
                graph_name = '%s_numeric' %(column)
            else:
                graph_name = column

            if '/' in graph_name:
                graph_name = graph_name.replace('/', '')

            img = openpyxl.drawing.image.Image(os.path.join(img_dir, '%s.png' %(graph_name)))
            ws.add_image(img, 'C%d' %(head_row+1))
        except:
            continue
    # adjust worksheet
    _adjust_ws(ws=ws, row_height=row_height)
    ws.column_dimensions['C'].width = 90
