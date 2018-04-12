import pandas as pd
import numpy as np
import os
import shutil

import openpyxl
from sklearn.externals.joblib import Parallel, delayed

import matplotlib.pyplot as plt
import seaborn as sns 
sns.set_style('white')

import datetime
from scipy.stats import spearmanr

import warnings
warnings.filterwarnings('ignore')

from .dqc_utils import _get_scale_draw_values, _draw_texts, _insert_numeric_results
from .data_compare import _compare_key, _check_features, _insert_summary
from .data_summary import _insert_string_results

# global color values
TABLE1_DARK = "#4BACC6"
TABLE1_LIGHT = "#DAEEF3"

TABLE2_DARK = "#F79646"
TABLE2_LIGHT = "#FDE9D9"


def numeric_consist_pretty(_df1, _df2, _key1, _key2, col, figsize=None, date_flag=False):
    """
    Draw pretty distribution graph for checking data consistency

    Parameters
    ----------
    _df1: pandas DataFrame
        slice of table1 containing enough information to check
    _df2: pandas DataFrame
        slice of table2 containing enough information to check
    _key1: string
        key for table1
    _key2: string
        key for table2
    col: string
        name of column to check
    figsize: tuple, default=None
        figure size
    date_flag: bool, default=False
        whether it is checking date features
    """

    # color values for graph
    TABLE1_DARK = "#4BACC6"
    TABLE2_DARK = "#F79646"

    df = _df1.merge(_df2, left_on=_key1, right_on=_key2, how="inner")
    df['diff_temp'] = df['%s_y' %(col)] - df['%s_x' %(col)]
    draw_values = df['diff_temp'].dropna().values
    origin_value_4 = [np.min(draw_values), np.mean(draw_values), np.median(draw_values), np.max(draw_values)]

    # get distribution
    scale_flg = 0
    draw_value_4 = origin_value_4
    if np.max([abs(origin_value_4[0]), abs(origin_value_4[3])]) >= pow(10, 6):
        scale_flg = 1
        draw_values, draw_value_4 = _get_scale_draw_values(draw_values, draw_value_4)

    plt.clf()
    if figsize is not None:
        plt.figure(figsize)
    else:
        plt.figure(figsize=(9, 4))

    both_min = np.min([df['%s_x' %(col)].min(), df['%s_y' %(col)].min()])
    both_max = np.max([df['%s_x' %(col)].max(), df['%s_y' %(col)].max()])

    plt.subplot(121)
    plt.title('Scatter plot for values')
    plt.scatter(df['%s_x' %(col)].values, df['%s_y' %(col)].values, c=TABLE1_DARK, s=5)
    plt.plot([both_min, both_max], [both_min, both_max], '--', c='#bbbbbb')

    plt.xlim(both_min, both_max)
    plt.ylim(both_min, both_max)

    ax2 = plt.subplot(122)
    if len(np.unique(draw_values)) <= 10:
        sns.countplot(draw_values, palette=sns.color_palette([TABLE2_DARK]))
        if len(np.unique(draw_values)) > 5:
            plt.xticks(rotation=90)
    else:
        sns.distplot(draw_values, color=TABLE2_DARK)
        y_low, y_up = ax2.get_ylim()
        _draw_texts(text_values=origin_value_4, draw_value_4=draw_value_4, mark=1, y_low=y_low, y_up=y_up)

    if date_flag:
        plt.title('Distribution of differences (in months)')
    elif scale_flg:
        plt.title('Distribution of differences (log10 scale)')
    else:
        plt.title('Distribution of differences')

    plt.show()


def _consist_numeric(col, _df1, _df2, _key1, _key2, img_dir, date_flag=False):
    """
    Check consistency for numeric type column

    Parameters
    ----------
    col: string
        name of column to check
    _df1: pandas DataFrame
        slice of table1 containing enough information to check
    _df2: pandas DataFrame
        slice of table2 containing enough information to check
    _key1: column to merge on for table1
    _key2: column to merge on for table2
    img_dir: root directory for the generated images
    date_flag: boolean
        Whether the column is date type

    Returns
    -------
    Dictionary contains the output result
    """

    df1, df2 = _df1.copy(), _df2.copy()
    df = pd.merge(df1, df2, left_on=_key1, right_on=_key2, how="inner")

    if (df['%s_x' %(col)].dropna().shape[0] == 0) or (df['%s_y' %(col)].dropna().shape[0] == 0):
        if (df['%s_x' %(col)].dropna().shape[0] == 0) and (df['%s_y' %(col)].dropna().shape[0] == 0):
            error_msg = 'all nan in both table'
        elif df['%s_x' %(col)].dropna().shape[0] == 0:
            error_msg = 'all nan in table1'
        else:
            error_msg = 'all nan in table2'
        return {'column': col, 'error_msg': error_msg}

    df = df.dropna(how='any', subset=['%s_x' % (col), '%s_y' % (col)]).reset_index(drop=True)
    df['diff_temp'] = df['%s_y' %(col)] - df['%s_x' %(col)]
    corr = round(spearmanr(df['%s_x' %(col)].values, df['%s_y' %(col)].values)[0], 3)

    output = [
        {'feature': 'column', 'value': col, 'graph': 'consistency check'},
        {'feature': 'corr', 'value': corr},
        {'feature': 'min diff', 'value': round(df['diff_temp'].min(), 3)},
        {'feature': 'mean diff', 'value': round(df['diff_temp'].mean(), 3)},
        {'feature': 'median diff', 'value': round(df['diff_temp'].median(), 3)},
        {'feature': 'max diff', 'value': round(df['diff_temp'].max(), 3)},
    ]

    draw_values = df['diff_temp'].dropna().values
    origin_value_4 = [np.min(draw_values), np.mean(draw_values), np.median(draw_values), np.max(draw_values)]

    # get distribution
    scale_flg = 0
    draw_value_4 = origin_value_4
    if np.max([abs(origin_value_4[0]), abs(origin_value_4[3])]) >= pow(10, 6):
        scale_flg = 1
        draw_values, draw_value_4 = _get_scale_draw_values(draw_values, draw_value_4)

    # draw the scatter plot
    both_min = np.min([df['%s_x' %(col)].min(), df['%s_y' %(col)].min()])
    both_max = np.max([df['%s_x' %(col)].max(), df['%s_y' %(col)].max()])

    dpi = 72
    plt.figure(figsize=(635. / dpi, 635. / (9. / 4.) / dpi), dpi=dpi)
    plt.subplot(121)
    plt.title('Scatter plot for values')
    plt.scatter(df['%s_x' %(col)].values, df['%s_y' %(col)].values, c=TABLE1_DARK, s=5)
    plt.plot([both_min, both_max], [both_min, both_max], '--', c='#bbbbbb')

    plt.xlim(both_min, both_max)
    plt.ylim(both_min, both_max)

    ax2 = plt.subplot(122)
    if len(np.unique(draw_values)) <= 10:
        sns.countplot(draw_values, palette=sns.color_palette([TABLE2_DARK]))
        if len(np.unique(draw_values)) > 5:
            plt.xticks(rotation=90)
    else:
        sns.distplot(draw_values, color=TABLE2_DARK)
        y_low, y_up = ax2.get_ylim()
        _draw_texts(text_values=origin_value_4, draw_value_4=draw_value_4, mark=1, y_low=y_low, y_up=y_up)

    if date_flag:
        plt.title('Distribution of differences (in months)')
    elif scale_flg:
        plt.title('Distribution of differences (log10 scale)')
    else:
        plt.title('Distribution of differences')

    # save the graphs
    # adjust graph name
    graph_name = col
    if '/' in graph_name:
        graph_name = graph_name.replace('/', '')
    plt.savefig(os.path.join(img_dir, graph_name + '.png'), transparent=True, dpi=dpi)
    return {'column': col, 'result_df': pd.DataFrame(output), 'corr': {'column': col, 'corr': corr}}


def _consist_string(col, _df1, _df2, _key1, _key2):
    """
    Check consistency for string type column

    Parameters
    ----------
    col: string
        name of column to check
    _df1: pandas DataFrame
        slice of table1 containing enough information to check
    _df2: pandas DataFrame
        slice of table2 containing enough information to check
    _key1: column to merge on for table1
    _key2: column to merge on for table2

    Returns
    -------
    Dictionary contains the output result
    """

    df1, df2 = _df1.copy(), _df2.copy()
    df = pd.merge(df1, df2, left_on=_key1, right_on=_key2, how="inner")

    if (df['%s_x' %(col)].dropna().shape[0] == 0) or (df['%s_y' %(col)].dropna().shape[0] == 0):
        if (df['%s_x' %(col)].dropna().shape[0] == 0) and (df['%s_y' %(col)].dropna().shape[0] == 0):
            error_msg = 'all nan in both table'
        elif df['%s_x' %(col)].dropna().shape[0] == 0:
            error_msg = 'all nan in table1'
        else:
            error_msg = 'all nan in table2'
        return {'column': col, 'error_msg': error_msg}

    df['diff_temp'] = df.apply(lambda x: "Same" if x['%s_x' %(col)] == x['%s_y' %(col)] else "Diff", axis=1)
    df['diff_temp'] = df.apply(lambda x: "Same" if (str(x['%s_x' % (col)]) == 'nan'
                                                    and str(x['%s_y' % (col)]) == 'nan') else x['diff_temp'], axis=1)

    corr = round(df[df['diff_temp'] == "Same"].shape[0] * 1.0 / df.shape[0], 3)
    output = [
        {'feature': 'column', 'value': col},
        {'feature': 'corr', 'value': corr}
    ]

    if corr == 1:
        return {'column': col, 'result_df': [pd.DataFrame(output), pd.DataFrame()], 'corr': {'column': col, 'corr': corr}}
    else:
        diff_df = df[df['diff_temp'] == "Diff"].reset_index(drop=True)
        diff_df['diff_combo'] = diff_df['%s_x' %(col)].map(str) + ' -> ' + diff_df['%s_y' %(col)].map(str)
        diff_df_vc = pd.DataFrame(diff_df['diff_combo'].value_counts())
        diff_df_vc.columns = ['count']
        diff_df_vc['diff_combo'] = diff_df_vc.index.values
        diff_df_vc = diff_df_vc.sort_values(by='count', ascending=False).head(10)[['diff_combo', 'count']]

        return {'column': col, 'result_df': [pd.DataFrame(output), diff_df_vc],
                'corr': {'column': col, 'corr': corr}}


def data_consist(table1, table2, key1, key2, schema1, schema2, fname, sample_size=1.0,
                 output_root='', keep_images=False, n_jobs=1):
    """
    Check consistency between two tables

    Parameters
    ----------
    table1: pandas DataFrame
        one of the two tables to compare
    table2: pandas DataFrame
        one of the two tables to compare
    key1: string
        key for table1
    key2: string
        key for table2
    schema1: pandas DataFrame
        data schema (contains column names and corresponding data types) for _table1
    schema2: pandas DataFrame
        data schema (contains column names and corresponding data types) for _table2
    fname: string
        the output file name
    sample_size: integer or float(<=1.0), default=1.0
        int: number of sample rows to do the comparison (useful for large tables)
        float: sample size in percentage
    output_root: string, default=''
        the root directory for the output file
    keep_images: boolean, default=False
        whether to keep all generated images
    n_jobs: int, default=1
        the number of jobs to run in parallel
    """

    # check whether keys are valid
    if key1 not in table1.columns.values:
        raise ValueError('key1: does not exist in table1')
    if key2 not in table2.columns.values:
        raise ValueError('key2: does not exist in table2')

    # check whether two tables are unique in key level
    if table1[key1].nunique() != table1.shape[0]:
        raise ValueError('table1: should be unique in %s level' % (key1))
    if table2[key2].nunique() != table2.shape[0]:
        raise ValueError('table2: should be unique in %s level' % (key2))

    # check sample_size
    if sample_size > 1:
        if int(sample_size) != sample_size:
            raise ValueError('sample_size: only accept integer when it is > 1.0')
        if (sample_size > table1.shape[0]) or (sample_size > table2.shape[0]):
            print('sample_size: %d is smaller than %d or %d...' % (sample_size, table1.shape[0], table2.shape[0]))

    # check output_root
    if output_root != '':
        if not os.path.isdir(output_root):
            raise ValueError('output_root: root not exists')

    # create a new workbook to store everything
    wb = openpyxl.Workbook()

    # prepare directory for generated images
    img_dir = 'img_temp'
    if os.path.isdir(img_dir):
        shutil.rmtree(img_dir)
    os.mkdir(img_dir)

    # calculate the sample size
    if sample_size <= 1.0:
        both_keys = list(set(table1[key1].values).intersection(set(table2[key2].values)))
        sample_size = np.min([int(table1.shape[0] * sample_size), int(table2.shape[0] * sample_size), len(both_keys)])
        sample_keys = np.random.choice(both_keys, sample_size, replace=False)
        table1 = table1[table1[key1].isin(sample_keys)].reset_index(drop=True)
        table2 = table2[table2[key2].isin(sample_keys)].reset_index(drop=True)

    schema, check_features = _check_features(schema1, schema2)
    corr_results = []

    # key features
    key_features = check_features['key']
    if len(key_features) > 0:
        _n_jobs = np.min([n_jobs, len(key_features)])
        key_results = Parallel(n_jobs=_n_jobs)(delayed(_compare_key)(col, table1[[col]], table2[[col]], img_dir)
            for col in key_features)

        for key_result in key_results:
            if 'corr' in key_result.keys():
                corr_results.append(key_result['corr'])

        # write all results to worksheet
        ws = wb.create_sheet(title=u'key')
        _insert_numeric_results(key_results, ws, 40, img_dir)

    # numeric features
    numeric_features = check_features['numeric']
    if len(numeric_features) > 0:
        _n_jobs = np.min([n_jobs, len(numeric_features)])
        numeric_results = Parallel(n_jobs=_n_jobs)(delayed(_consist_numeric)(col, table1[[key1, col]],
            table2[[key2, col]], key1, key2, img_dir) for col in numeric_features)

        for numeric_result in numeric_results:
            if 'corr' in numeric_result.keys():
                corr_results.append(numeric_result['corr'])

        # write all results to worksheet
        ws = wb.create_sheet(title=u'numeric')
        _insert_numeric_results(numeric_results, ws, 45, img_dir)

    # string features
    string_features = check_features['str']
    if len(string_features) > 0:
        _n_jobs = np.min([n_jobs, len(string_features)])
        string_results = Parallel(n_jobs=_n_jobs)(delayed(_consist_string)(col, table1[[key1, col]],
            table2[[key2, col]], key1, key2) for col in string_features)

        for string_result in string_results:
            if 'corr' in string_result.keys():
                corr_results.append(string_result['corr'])

        # write all results to worksheet
        ws = wb.create_sheet(title=u'string')
        _insert_string_results(string_results, ws, 25)

    # date features
    date_features = check_features['date']
    if len(date_features) > 0:
        # get the current time
        snapshot_date_now = str(datetime.datetime.now().date())
        for col in date_features:
            table1[col] = (pd.to_datetime(snapshot_date_now) - pd.to_datetime(table1[col],
                errors='coerce')).astype('timedelta64[M]', errors='ignore')
            table2[col] = (pd.to_datetime(snapshot_date_now) - pd.to_datetime(table2[col],
                errors='coerce')).astype('timedelta64[M]', errors='ignore')
        _n_jobs = np.min([n_jobs, len(date_features)])
        date_results = Parallel(n_jobs=_n_jobs)(delayed(_consist_numeric)(col, table1[[key1, col]], table2[[key2, col]],
            key1, key2, img_dir, date_flag=True) for col in date_features)

        for date_result in date_results:
            if 'corr' in date_result.keys():
                corr_results.append(date_result['corr'])

        # write all results to worksheet
        ws = wb.create_sheet(title=u'date')
        _insert_numeric_results(date_results, ws, 45, img_dir, date_flag=True)

    # insert the summary
    _insert_summary(wb, schema, corr_results)

    wb.save(filename=os.path.join(output_root, 'data_consist_%s.xlsx' %(fname)))
    if not keep_images:
        shutil.rmtree(img_dir)


def data_consist_notebook(table1, table2, key1, key2, schema1, schema2, fname, output_root=''):
    """
    Automatically generate ipynb for checking data consistency

    Parameters
    ----------
    table1: pandas DataFrame
        one of the two tables to compare
    table2: pandas DataFrame
        one of the two tables to compare
    key1: string
        key for table1
    key2: string
        key for table2
    schema1: pandas DataFrame
        data schema (contains column names and corresponding data types) for _table1
    schema2: pandas DataFrame
        data schema (contains column names and corresponding data types) for _table2
    fname: string
        the output file name
    output_root: string, default=''
        the root directory for the output file
    """

    # check whether keys are valid
    if key1 not in table1.columns.values:
        raise ValueError('key1: does not exist in table1')
    if key2 not in table2.columns.values:
        raise ValueError('key2: does not exist in table2')

    # check whether two tables are unique in key level
    if table1[key1].nunique() != table1.shape[0]:
        raise ValueError('table1: should be unique in %s level' % (key1))
    if table2[key2].nunique() != table2.shape[0]:
        raise ValueError('table2: should be unique in %s level' % (key2))

    # check output_root
    if output_root != '':
        if not os.path.isdir(output_root):
            raise ValueError('output_root: root not exists')

    # generate output file path
    output_path = os.path.join(output_root, 'data_consist_notebook_%s.py' %(fname))

    # delete potential generated script and notebook
    if os.path.isfile(output_path):
        os.remove(output_path)

    if os.path.isfile(output_path.replace('.py', '.ipynb')):
        os.remove(output_path.replace('.py', '.ipynb'))

    schema, _ = _check_features(schema1, schema2)

    dir_path = os.path.dirname(os.path.realpath(__file__))
    main_line = open(dir_path + '/templates/data_consist_main.txt').read()
    key_line = open(dir_path + '/templates/data_consist_key.txt').read()
    str_line = open(dir_path + '/templates/data_consist_str.txt').read()
    numeric_line = open(dir_path + '/templates/data_consist_numeric.txt').read()
    date_line = open(dir_path + '/templates/data_consist_date.txt').read()

    with open(output_path, "a") as outbook:
        outbook.write(main_line)

        schema_error = schema[schema['error'] != ''].reset_index(drop=True)
        if schema_error.shape[0] > 0:
            schema_error['error_column'] = schema_error['column_1']
            schema_error.loc[schema_error['error_column'].isnull(), 'error_column'] = schema_error['column_2']
            outbook.write('\n"""\n## error columns\n\n')
            for i in range(schema_error.shape[0]):
                get = schema_error.iloc[i]
                outbook.write('**%s:** %s<br>' % (get['error_column'], get['error']))
            outbook.write('"""\n\n')

        # only compare check columns in both table1 and table2, and follow the column order of table1
        check_cols = [col for col in table1.columns.values if col in schema[schema['error'] == '']['column_1'].values]
        for col in check_cols:
            # get the data type of the column
            col_type = schema[schema['column_1'] == col]['type_1'].values[0]

            outbook.write('\n"""\n## %s (type: %s)\n\n"""\n\n' %(col, col_type))
            outbook.write('col = "%s"\n' %(col))

            # for key and str, compare intersection
            if col_type == 'key':
                outbook.write(key_line)
            elif col_type == 'str':
                outbook.write(str_line)
            elif col_type == 'numeric':
                outbook.write(numeric_line)
            else:
                outbook.write(date_line)
        outbook.close()

    os.system("python -m py2nb %s %s" %(output_path, output_path.replace('.py', '.ipynb')))